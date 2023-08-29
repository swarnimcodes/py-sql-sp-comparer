import os
import pandas as pd
import filecmp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# Define the base directory and source/test directories
# base_dir = "C:\\Users\\swarn\\github\\Stored SPs"
base_dir = "C:\\Users\\swarn\\Downloads\\NEW PHINMA UAT SCRIPTS"
source_dir = os.path.join(base_dir, "SJC")

test_dirs = [
    os.path.join(base_dir, "SJC"),
    os.path.join(base_dir, "HK"),
    os.path.join(base_dir, "COC"),
    os.path.join(base_dir, "RC"),
]

# Create a list to store the data for each SP
sp_data = []

# Get the list of SQL files in the source directory
source_sql_files = os.listdir(source_dir)

def strip_comments(sql_file_contents):
    stripped_sql_file = re.sub(r'--.*', '', sql_file_contents)
    stripped_sql_file = re.sub(r'/\*.*?\*/', '', stripped_sql_file, flags=re.DOTALL)
    stripped_sql_file = '\n'.join(' '.join(part.strip() for part in line.split() if part.strip()) for line in stripped_sql_file.splitlines() if line.strip())
    return stripped_sql_file

def difference(source_sql_path, test_sql_path) -> bool:
    source_contents = open(source_sql_path, 'r').read().upper()
    test_contents = open(test_sql_path, 'r').read().upper()

    stripped_sql_file_source = strip_comments(source_contents)
    stripped_sql_file_test = strip_comments(test_contents)

    if stripped_sql_file_source == stripped_sql_file_test:
        return True
    else:
        return False

# Iterate through each SQL file in the source directory
for sql_file in source_sql_files:
    sp_name = sql_file[:-4]  # Remove the ".sql" extension

    # Create a dictionary to store the presence in different test directories
    sp_info = {'SP Name': sp_name}

    source_sql_path = os.path.join(source_dir, sql_file)

    # Check presence in each test directory
    for test_dir in test_dirs:
        test_sql_path = os.path.join(test_dir, sql_file)
        if os.path.exists(test_sql_path):
            if difference(source_sql_path, test_sql_path):
                sp_info[os.path.basename(test_dir)] = 'PRESENT & EQUAL'
            else:
                sp_info[os.path.basename(test_dir)] = 'PRESENT & UNEQUAL'
        else:
            sp_info[os.path.basename(test_dir)] = 'ABSENT'

    sp_data.append(sp_info)

# Create a DataFrame from the collected data
df = pd.DataFrame(sp_data)

# Define the path for the output Excel file
output_excel_path = os.path.join(base_dir, 'SP_Presence_Report_PHINMA_UAT.xlsx')

# Write the DataFrame to an Excel file
df.to_excel(output_excel_path, index=False)

# Load the existing workbook and sheet
wb = load_workbook(output_excel_path)
ws = wb.active

# Apply cell coloring based on the cell values
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value == 'PRESENT & UNEQUAL':
            cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
        elif cell.value == 'ABSENT':
            cell.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")

# Save the modified workbook
wb.save(output_excel_path)
