import nltk
import os
import difflib
import sqlparse
import re
import pandas as pd
from joblib import Parallel, delayed, Memory
import multiprocessing
import hashlib
from openpyxl import Workbook
from openpyxl.styles import PatternFill

cachedir = './cache'  # Directory to store cached results
memory = Memory(cachedir, verbose=0)

output_dir = "diffs"
excel_output_file = "SP_Comparison.xlsx"

# Source Database
folder1_path = r"C:\\Users\\swarn\\github\\Stored SPs\\DB_PRMITR_ERP_20230701"

# Database to compare with source
folder2_path = r"C:\\Users\\swarn\\github\\Stored SPs\\DB_PRMITR_ERP_20230615"

def strip_sql_comments(sql_file_contents) -> str:
    stripped_sql_file = re.sub(r'--.*', '', sql_file_contents)
    stripped_sql_file = re.sub(r'/\*.*?\*/', '', stripped_sql_file, flags=re.DOTALL)
    stripped_sql_file = '\n'.join(' '.join(part.strip() for part in line.split() if part.strip()) for line in stripped_sql_file.splitlines() if line.strip())
    return stripped_sql_file

def normalize_sql(file_contents):
    file_contents = file_contents
    tokens = nltk.word_tokenize(file_contents)
    return tokens

def generate_html_diff(file1_contents, file2_contents, folder1_path, folder2_path):
    folder1_name = os.path.basename(folder1_path)
    folder2_name = os.path.basename(folder2_path)
    file1_formatted = sqlparse.format(file1_contents, reindent=True, keyword_case='upper', use_space_around_operators=True, indent_width=4).strip()
    file2_formatted = sqlparse.format(file2_contents, reindent=True, keyword_case='upper', use_space_around_operators=True, indent_width=4).strip()

    html_diff = difflib.HtmlDiff(tabsize=4, wrapcolumn=72).make_file(file1_formatted.splitlines(), file2_formatted.splitlines(), context=True, numlines=1, charset='utf-8', fromdesc=folder1_name, todesc=folder2_name)
    return html_diff

# A dictionary to store cached hashes
hash_cache = {}

def get_file_hash(file_path):
    if file_path in hash_cache:
        return hash_cache[file_path]
    else:
        file_contents = open(file_path, 'r').read().upper()
        file_nocomments = strip_sql_comments(file_contents)
        normalized_sql = normalize_sql(file_nocomments)
        normalized_sql_str = ' '.join(normalized_sql)
        file_hash = hashlib.sha1(normalized_sql_str.encode()).hexdigest()
        hash_cache[file_path] = file_hash
        return file_hash

def process_sql_file(sql_file, folder1_path, folder2_path):
    file1_path = os.path.join(folder1_path, sql_file)
    file2_path = os.path.join(folder2_path, sql_file)

    sp_name = sql_file

    present_in_folder1 = os.path.exists(file1_path)
    present_in_folder2 = os.path.exists(file2_path)
    content_comparison = ""

    if present_in_folder1 and present_in_folder2:
        file1_contents = open(file1_path, 'r').read().upper()
        file2_contents = open(file2_path, 'r').read().upper()

        file1_nocomments = strip_sql_comments(file1_contents)
        normalized_sql_1 = normalize_sql(file1_nocomments)

        file2_nocomments = strip_sql_comments(file2_contents)
        normalized_sql_2 = normalize_sql(file2_nocomments)

        if normalized_sql_1 == normalized_sql_2:
            content_comparison = "Equal"
            print(f"Files {sql_file} are equal")
        else:
            content_comparison = "Different"
            print(f"Files {sql_file} are unequal")
            folder1_name = os.path.basename(folder1_path)
            folder2_name = os.path.basename(folder2_path)
            diff_html = generate_html_diff(file1_nocomments, file2_nocomments, folder1_name, folder2_name)
            diff_filename = os.path.join(output_dir, f"diff_{sql_file}.html")
            with open(diff_filename, "w") as diff_file:
                diff_file.write(diff_html)
    else:
        content_comparison = "Missing in one of the folders"

    return [sp_name, present_in_folder1, present_in_folder2, content_comparison]

def main():
    num_cores = multiprocessing.cpu_count()
    comparison_results = Parallel(n_jobs=num_cores)(
        delayed(process_sql_file)(sql_file, folder1_path, folder2_path)
        for sql_file in os.listdir(folder1_path) if sql_file.endswith(".sql")
    )

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Set column headers
    ws.append(["SP Name", "Present in DB_PRMITR_ERP_20230701", "Present in DB_PRMITR_ERP_20230615", "Content Comparison"])

    # Define colors
    missing_color = PatternFill(start_color="FFD5B5", end_color="FFD5B5", fill_type="solid")
    different_color = PatternFill(start_color="FF99FF", end_color="FF99FF", fill_type="solid")

    # Populate Excel rows and apply cell coloring
    for result in comparison_results:
        row = [result[0], result[1], result[2], result[3]]
        ws.append(row)

        excel_row = ws[ws.max_row]

        if not result[1]:
            excel_row[1].fill = missing_color

        if not result[2]:
            excel_row[2].fill = missing_color

        if result[3] == "Different":
            excel_row[3].fill = different_color


    # Save the Excel file
    wb.save(excel_output_file)

if __name__ == "__main__":
    main()
