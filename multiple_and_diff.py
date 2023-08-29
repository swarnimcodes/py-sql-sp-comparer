import os
import pandas as pd
import filecmp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import difflib
import nltk
import sqlparse

### UPLOAD TO GIT

# Define the base directory and source/test directories
# base_dir = "C:\\Users\\swarn\\github\\Stored SPs"
# source_dir = os.path.join(base_dir, "DB_PRMITR_ERP_20230701")
# test_dirs = [
#     os.path.join(base_dir, "DB_PRMITR_ERP_20230615_TEST_1"),
#     os.path.join(base_dir, "DB_PRMITR_ERP_20230615_TEST_2"),
#     os.path.join(base_dir, "DB_PRMITR_ERP_20230615_TEST_3"),
#     os.path.join(base_dir, "DB_PRMITR_ERP_20230701")
# ]

base_dir = "C:\\Users\\swarn\\Downloads\\PHINMA UAT"
source_dir = os.path.join(base_dir, "HK")
diff_op_dir = "C:\\Users\\swarn\\Downloads\\PHINMA UAT\\diffs"


test_dirs = [
    os.path.join(base_dir, "HK"),
    os.path.join(base_dir, "COC"),
    os.path.join(base_dir, "RC"),
    os.path.join(base_dir, "SJC")
]


# Create a list to store the data for each SP
sp_data = []

# Get the list of SQL files in the source directory
source_sql_files = os.listdir(source_dir)

# def strip_comments(sql_file_contents):
#     stripped_sql_file = sql_file_contents.strip()
#     stripped_sql_file_normalized = normalize_sql(stripped_sql_file)

#     stripped_sql_file = re.sub(r'--.*', '', sql_file_contents)
#     stripped_sql_file = re.sub(r'/\*.*?\*/', '', sql_file_contents, flags=re.DOTALL)
#     stripped_sql_file = '\n'.join(''.join(part.strip() for part in line.split() if part.strip()) for line in sql_file_contents.splitlines() if line.strip())
#     stripped_sql_file_normalized = normalize_sql(stripped_sql_file)

#     stripped_sql_file_formatted = sqlparse.format(stripped_sql_file_normalized, reindent=True, keyword_case='upper', indent_width=4).strip()

#     return stripped_sql_file_formatted

def strip_sql_comments(sql_file_contents) -> str:
    stripped_sql_file = re.sub(r'--.*', '', sql_file_contents)
    stripped_sql_file = re.sub(r'/\*.*?\*/', '', stripped_sql_file, flags=re.DOTALL)
    stripped_sql_file = '\n'.join(' '.join(part.strip() for part in line.split() if part.strip()) for line in stripped_sql_file.splitlines() if line.strip())
    return stripped_sql_file

def normalize_sql(file_contents): ## CHANGE
    file_contents = file_contents
    tokens = nltk.word_tokenize(file_contents)
    print(tokens)
    return tokens

# def normalize_sql(file_contents): ## CHANGE
#     file_contents = file_contents
#     tokens = nltk.word_tokenize(file_contents)
#     normalized_sql_str = ''.join(tokens)
#     return normalized_sql_str

def difference(source_sql_path, test_sql_path) -> bool:
    source_contents = open(source_sql_path, 'r').read().upper()
    test_contents = open(test_sql_path, 'r').read().upper()

    stripped_sql_file_source = normalize_sql(source_contents)
    stripped_sql_file_source = strip_sql_comments(stripped_sql_file_source)

    stripped_sql_file_test = normalize_sql(test_contents)
    stripped_sql_file_test = strip_sql_comments(stripped_sql_file_test)

    if (stripped_sql_file_source == stripped_sql_file_test):
        return True
    else:
        return False

# def generate_html_diff(file1_contents, file2_contents, folder1_path, folder2_path):
#     folder1_name = os.path.basename(folder1_path)
#     folder2_name = os.path.basename(folder2_path)
#     file1_formatted = sqlparse.format(file1_contents, reindent=True, keyword_case='upper', use_space_around_operators=True, indent_width=4).strip()
#     file2_formatted = sqlparse.format(file2_contents, reindent=True, keyword_case='upper', use_space_around_operators=True, indent_width=4).strip()

#     html_diff = difflib.HtmlDiff(tabsize=4, wrapcolumn=72).make_file(file1_formatted.splitlines(), file2_formatted.splitlines(), context=True, numlines=1, charset='utf-8', fromdesc=folder1_name, todesc=folder2_name)
#     return html_diff

def gen_diff(source_sql_path, test_sql_path):
    source_contents = open(source_sql_path, 'r').read().upper()
    test_contents = open(test_sql_path, 'r').read().upper()

    stripped_sql_file_source = strip_sql_comments(source_contents)
    stripped_sql_file_test = strip_sql_comments(test_contents)

    normalized_source_file_contents = normalize_sql(stripped_sql_file_source)
    normalized_test_file_contents = normalize_sql(stripped_sql_file_test)

    normalized_source_file = ''.join(normalized_source_file_contents)
    normalized_test_file = ''.join(normalized_test_file_contents)

    source_file_formatted = sqlparse.format(normalized_source_file, reindent=True, keyword_case='upper').strip()
    test_file_formatted = sqlparse.format(normalized_test_file, reindent=True, keyword_case='upper').strip()

    # print(source_file_formatted) # this is good.... then why html file has spaces???

    html_diff = difflib.HtmlDiff(tabsize=4, wrapcolumn=72).make_file(source_file_formatted.splitlines(), test_file_formatted.splitlines(), fromdesc=source_sql_path, todesc=test_sql_path)

    return html_diff

def main():
    # Iterate through each SQL file in the source directory
    for sql_file in source_sql_files:
        sp_name = sql_file[:-4]  # Remove the ".sql" extension

        # Create a dictionary to store the presence in different test directories
        sp_info = {'SP Name': sp_name}

        source_sql_path = os.path.join(source_dir, sql_file)

        # Check presence in each test directory
        for test_dir in test_dirs:
            test_sql_path = os.path.join(test_dir, sql_file)
            if os.path.exists(test_sql_path):
                if difference(source_sql_path, test_sql_path):
                    sp_info[os.path.basename(test_dir)] = 'present & equal'
                else:
                    sp_info[os.path.basename(test_dir)] = 'present & unequal' # only generate diff here
                    htmldiff = gen_diff(source_sql_path, test_sql_path)
                    # Define the diff filename within the diff directory
                    diff_filename = os.path.basename(test_sql_path) + "_diff.html"
                    diff_path = os.path.join(diff_op_dir, test_dir, diff_filename)

                    # Write the HTML diff to the specified file
                    with open(diff_path, 'w') as diff_file:
                        diff_file.write(htmldiff)
            else:
                sp_info[os.path.basename(test_dir)] = 'absent'

        sp_data.append(sp_info)

    # Create a DataFrame from the collected data
    df = pd.DataFrame(sp_data)

    # Define the path for the output Excel file
    output_excel_path = os.path.join(base_dir, 'SP_Presence_Report.xlsx')

    # Write the DataFrame to an Excel file
    df.to_excel(output_excel_path, index=False)

    # Load the existing workbook and sheet
    wb = load_workbook(output_excel_path)
    ws = wb.active

    # Apply cell coloring based on the cell values
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value == 'present & unequal':
                cell.fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
            elif cell.value == 'absent':
                cell.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")

    # Save the modified workbook
    wb.save(output_excel_path)

if __name__ == "__main__":
    main()
